from __future__ import annotations

from datetime import datetime, UTC

from aiogram import types

from classes.mongo_db_object import MongoDBObject
from classes.search_result import SearchResult

from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


class MultiQuery(MongoDBObject):
    collection_name = 'Queries'
    fields = {'_id': int, 'from_user_id': int, 'texts': list[str], 'datetime': datetime, "type": str}
    def __init__(self, _id: int,
                 from_user_id: int,
                 texts: list[str],
                 _datetime: datetime,
                 type_: str = 'multi'):
        self.__id = _id
        self.__from_user_id = from_user_id
        self.__texts = texts
        self.__datetime = _datetime
        self.__type = type_

    @property
    def from_user_id(self) -> int:
        return self.__from_user_id

    @property
    def _id(self) -> int:
        return self.__id

    @property
    def texts(self) -> list[str]:
        return self.__texts

    @property
    def type(self) -> str:
        return self.__type

    @property
    def datetime(self) -> datetime:
        return self.__datetime

    async def get_results(self) -> list[SearchResult]:
        return await SearchResult.get_many(self.texts)

    async def get_results_in_dict(self) -> dict[str, SearchResult]:
        return await SearchResult.get_many_in_dict(self.texts)

    async def get_result_by_text_n(self, n: int) -> SearchResult:
        if n < 0 or n >= len(self.texts):
            raise IndexError(f"Wrong text index. correct index must be in: [0, {len(self.texts) - 1}]")
        return await SearchResult.get(self.texts[n])

    @classmethod
    def create(cls, from_user_id: int, query_texts: list[str], type_: str = 'multi') -> MultiQuery:
        query = cls(cls.get_max_id() + 1, from_user_id, query_texts, datetime.now(UTC), type_)
        cls.collection.insert_one(query.to_json())
        return cls.from_json(query.to_json())

    @classmethod
    def get_by_id(cls, _id: int) -> MultiQuery:
        return cls.from_json(cls.collection.find_one({"_id": _id}, cls.fields_keys))

    @classmethod
    def delete_all(cls) -> None:
        cls.collection.drop()

    @classmethod
    def get_max_id(cls) -> int:
        try:
            return list(cls.collection.find({}, ['_id']).sort('_id', -1))[0]['_id']
        except IndexError:
            return -1

    @classmethod
    def get_all(cls) -> list[MultiQuery]:
        res = []
        for q_dict in cls.collection.find({}):
            res.append(cls.from_json(q_dict))
        return res

    @property
    def general_message_keyboard(self):
        return types.InlineKeyboardMarkup(inline_keyboard=[[types.InlineKeyboardButton(text='Экспортировать все результаты в Excel', callback_data=f'EXPORT MQ XLSX {self._id}')]])

    async def export_results_to_excel(self):
        bold_font = Font(bold=True)
        all_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'), )
        l_r_border = Border(left=Side('thin'), right=Side('thin'))
        l_r_t_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'))
        l_r_b_border = Border(left=Side('thin'), right=Side('thin'), bottom=Side('thin'))
        wb = Workbook()
        wb.remove(wb.active)
        for text in self.texts:
            result = await SearchResult.get(text)
            wb.create_sheet(text)
            ws = wb[text]
            i = 1
            ws.column_dimensions['A'].width = 100
            ws.column_dimensions['B'].width = 100
            ws.column_dimensions['C'].width = 100
            for brand in result.brands_list:
                ws['A' + str(i)].value = brand.name
                ws['A' + str(i)].border = all_border
                ws['A' + str(i)].font = bold_font
                ws.merge_cells('A' + str(i) + ':C' + str(i))
                i += 1
                ws['A' + str(i)].value = 'Артикул'
                ws['B' + str(i)].value = 'Наименование'
                ws['C' + str(i)].value = 'Наличие'
                for letter in 'ABC':
                    ws[letter + str(i)].border = all_border
                    ws[letter + str(i)].font = bold_font
                i += 1
                for sp_stripped in result.spare_parts[brand.uid]:
                    sp = await sp_stripped.get_full_info()
                    ws['A' + str(i)].value = sp.code
                    ws['A' + str(i)].border = l_r_border
                    ws['B' + str(i)].value = sp.name
                    ws['B' + str(i)].border = l_r_border
                    counts_string = ''
                    if sp.counts:
                        counts_string += 'В наличии:'
                    else:
                        counts_string += 'Нет в наличии. '
                    for count in sp.counts:
                        counts_string += f"\n{count}"
                    ws.row_dimensions[i].height = 15 * (len(sp.counts) + 1)
                    ws['C' + str(i)].alignment = Alignment(wrapText=True)
                    ws['C' + str(i)].value = counts_string
                    ws['C' + str(i)].border = l_r_border
                    i += 1
                for letter in 'ABC':
                    ws[letter + str(i - 1)].border = l_r_b_border
                i += 1
        wb.save('result.xlsx')

if __name__ == '__main__':
    MultiQuery.delete_all()
    print(datetime.utcnow())
